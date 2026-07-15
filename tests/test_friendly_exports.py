import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from fund_ranking_system.friendly_exports import save_friendly_exports


class FriendlyExportsTest(unittest.TestCase):
    def test_save_friendly_exports_creates_office_files(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            reports_dir = root / "reports"
            processed_dir = root / "processed"
            reports_dir.mkdir()
            processed_dir.mkdir()
            (reports_dir / "fund_analysis_report.md").write_text(
                "# 主报告\n\n## 结论\n\n- 样例结论\n\n| 指标 | 数值 |\n|---|---:|\n| Score | 80 |",
                encoding="utf-8",
            )
            pd.DataFrame(
                {"fund": ["000001"], "rank": [1], "composite_score": [80.0]}
            ).to_csv(reports_dir / "ranking_balanced.csv", index=False)
            pd.DataFrame(
                {"fund": ["000001"], "annual_return": [0.1]}
            ).to_csv(processed_dir / "fund_metrics.csv", index=False)
            pd.DataFrame(
                {"portfolio": ["动态权重TopN等权"], "annual_return": [0.08]}
            ).to_csv(reports_dir / "portfolio_summary.csv", index=False)
            pd.DataFrame(
                {"portfolio": ["动态权重TopN等权"], "fund": ["000001"], "weight": [1.0]}
            ).to_csv(reports_dir / "portfolio_weights_balanced.csv", index=False)
            pd.DataFrame(
                {"constraint": ["组合目标"], "value": ["balanced"], "display_value": ["平衡组合"]}
            ).to_csv(reports_dir / "portfolio_constraints.csv", index=False)
            pd.DataFrame(
                {"fund": ["000001"], "selection_reason": ["样例原因"]}
            ).to_csv(reports_dir / "portfolio_recommendations.csv", index=False)
            pd.DataFrame(
                {"control_type": ["type_exposure"], "item": ["混合型"], "value": [1.0]}
            ).to_csv(reports_dir / "portfolio_risk_controls.csv", index=False)
            (reports_dir / "portfolio_recommendation.md").write_text("# 组合建议说明书\n\n样例", encoding="utf-8")

            docx_path, pdf_path, xlsx_path = save_friendly_exports(
                reports_dir,
                processed_dir,
                profile="balanced",
            )

            self.assertTrue(docx_path.exists())
            self.assertTrue(pdf_path.exists())
            self.assertTrue(xlsx_path.exists())
            self.assertGreater(docx_path.stat().st_size, 0)
            self.assertGreater(pdf_path.stat().st_size, 0)
            self.assertGreater(xlsx_path.stat().st_size, 0)
            workbook = load_workbook(xlsx_path, read_only=True)
            self.assertIn("组合构建摘要", workbook.sheetnames)
            self.assertIn("组合持仓权重", workbook.sheetnames)
            self.assertIn("组合约束配置", workbook.sheetnames)
            self.assertIn("组合建议明细", workbook.sheetnames)
            self.assertIn("组合风险控制", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
